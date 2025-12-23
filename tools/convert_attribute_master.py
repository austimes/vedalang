#!/usr/bin/env python3
"""
Convert Attribute-master-export.xlsx to attribute-master.json.

This is a one-off script to convert the VEDA attribute master export
into a JSON format suitable for runtime validation in table_schemas.py.

Usage:
    uv run python tools/convert_attribute_master.py

The output file is written to veda/attribute-master.json.
"""

import json
from pathlib import Path

import openpyxl


def convert_attribute_master():
    """Convert Excel attribute master to JSON format."""
    input_path = Path("veda/Attribute-master-export.xlsx")
    output_path = Path("vedalang/schema/attribute-master.json")

    wb = openpyxl.load_workbook(input_path, read_only=True)
    sheet = wb.active

    # Get header row
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Find column indices
    col_idx = {h: i for i, h in enumerate(headers) if h}

    attributes = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        attr_name = row[col_idx["Attribute"]]
        if not attr_name:
            continue

        # Parse aliases (comma-separated) - need this first for column_headers
        alias_str = row[col_idx.get("Alias", -1)] or ""
        aliases = []
        if alias_str:
            aliases = [a.strip() for a in alias_str.split(",") if a.strip()]

        # Build list of valid column headers (canonical + all aliases, lowercase)
        column_headers = [attr_name.lower()]
        column_headers.extend(a.lower() for a in aliases)

        # Build attribute entry
        entry = {
            "column_header": attr_name.lower(),  # Primary/canonical header
            "column_headers": column_headers,  # All valid headers (including aliases)
            "description": row[col_idx.get("Description", -1)] or "",
            "time_series": row[col_idx.get("TimeSeries", -1)] == "Yes",
            "process": row[col_idx.get("Process", -1)] == "T",
            "commodity": row[col_idx.get("Commodity", -1)] == "T",
            "timeslice": row[col_idx.get("TimeSlice", -1)] or "",
            "limtype": row[col_idx.get("LimType", -1)] or "",
            "currency": row[col_idx.get("Currency", -1)] == "CUR",
        }

        if aliases:
            entry["aliases"] = aliases

        # Parse indexes
        indexes_str = row[col_idx.get("Indexes", -1)] or ""
        if indexes_str:
            entry["indexes"] = indexes_str

        attributes[attr_name] = entry

    # Write JSON output
    output = {
        "_comment": "Auto-generated from Attribute-master-export.xlsx. Do not edit.",
        "_source": "tools/convert_attribute_master.py",
        "attributes": attributes,
    }

    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)

    print(f"Converted {len(attributes)} attributes to {output_path}")
    return attributes


if __name__ == "__main__":
    convert_attribute_master()
