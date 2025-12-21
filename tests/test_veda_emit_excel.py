import tempfile
from pathlib import Path

import jsonschema
import pytest
from openpyxl import load_workbook

from tools.veda_emit_excel import emit_excel, load_tableir, validate_tableir

PROJECT_ROOT = Path(__file__).parent.parent
EXAMPLES_DIR = PROJECT_ROOT / "vedalang" / "examples"


def test_emit_minimal_tableir():
    """Emit tableir_minimal.yaml and verify Excel structure."""
    tableir = load_tableir(EXAMPLES_DIR / "tableir_minimal.yaml")

    with tempfile.TemporaryDirectory() as tmpdir:
        created = emit_excel(tableir, Path(tmpdir))

        assert len(created) >= 1
        for path in created:
            assert path.exists()
            wb = load_workbook(path)
            assert len(wb.sheetnames) > 0


def test_excel_contains_tag():
    """Verify emitted Excel contains the table tag."""
    tableir = {
        "files": [
            {
                "path": "test.xlsx",
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [{"tag": "~FI_TEST", "rows": [{"COL1": "value1"}]}],
                    }
                ],
            }
        ]
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        created = emit_excel(tableir, Path(tmpdir))
        wb = load_workbook(created[0])
        ws = wb.active
        assert ws.cell(1, 1).value == "~FI_TEST"
        assert ws.cell(2, 1).value == "COL1"
        assert ws.cell(3, 1).value == "value1"


def test_invalid_tableir_rejected():
    """Invalid TableIR should raise ValidationError."""
    invalid = {
        "files": [
            {
                "path": "x.xlsx",
                "sheets": [
                    {"name": "S", "tables": [{"tag": "NO_TILDE", "rows": []}]}
                ],
            }
        ]
    }
    with pytest.raises(jsonschema.ValidationError):
        validate_tableir(invalid)
